from __future__ import unicode_literals
from django.http import HttpResponse
from django.template import loader
from django.shortcuts import render
from django.http import Http404
from django.shortcuts import get_object_or_404
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
from django.shortcuts import redirect
from django.contrib.auth import logout as django_logout
from django.contrib.auth.decorators import login_required

from .forms import *
from django.contrib.auth.models import User
import openpyxl
from  qa1.models import *
from readingmaterial.models import *
from django.shortcuts import render_to_response
from django.contrib import messages
from django.db.models import Q
from django.http import HttpResponseRedirect
from  django.template.context_processors import csrf

from django.contrib.auth.decorators import user_passes_test
from django.contrib.auth.models import Permission
from django.contrib.auth.decorators import permission_required

from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

from readingmaterial.models import *
from subscription.models import *
from qa1.models import *

import time
import random
from django.db.models import Q
from django.db import transaction
from django.db import IntegrityError


def getString(title):
    try:
        title = str(title)
    except Exception:
        title = title.encode('UTF8')

    # print ("\n ********** retirng strign: " + title)

    # print (title)
    return title



def create_mcq_from_excel(request, sheet, row, question_set):
    if ((sheet.cell(row=row, column=1).value) is None):
        # print ("in if part: " + str(sheet.cell(row=row, column=1).value))
        print ("************* NONE returning **********")
        return None
    # print ('***********.....mcq title is about to be been created \n')
    title = (sheet.cell(row=row, column=1).value)
    if (title):
        title = getString(title)
        title = title.decode('UTF8')
        q = Mcq_Question(question_text=title)
        q.question_set = question_set
    # print ('***********.....mcq title has been created \n')
    #q = Mcq_Question(question_text="alamin is hungry now")
    # print ("mcq created with title ")
    # print (q)
    # q.save()

    c1 = (sheet.cell(row=row, column=2).value)
    if (c1):
    	c1 = getString(c1)

    c2 = (sheet.cell(row=row, column=3).value)
    if(c2):
        c2 = getString(c2)

        # c2 = c2.encode('utf8')

    c3 = (sheet.cell(row=row, column=4).value)
    if (c3):
        c3 = getString(c3)

    c4 = (sheet.cell(row=row, column=5).value)
    if(c4):
        c4 = getString(c4)

    # c5 = (sheet.cell(row=row, column=8).value)
    # if (c5):
    #     c5= getString(c5)



    # c6 = (sheet.cell(row=row, column=9).value)
    # if (c6):

    #     try:
    #         c6 = str(c6)
    #     except Exception:
    #         c6 = c6.encode('utf8')


    reading_content = (sheet.cell(row=row, column=8).value)
    print (reading_content)
    if(reading_content):
        reading_content = getString(reading_content)
        reading_content = ReadingContent.objects.filter(id=int(reading_content)).first()

        print (" ******* Found REading content id *****")

    subtopic1 = (sheet.cell(row=row, column=9).value)
    if(subtopic1):
        subtopic1 = getString(subtopic1)
        subtopic1 = SubTopic1.objects.filter(id=int(subtopic1)).first()


    topic = (sheet.cell(row=row, column=10).value)
    if(topic):
        topic = getString(topic)
        topic = ReadingTopic.objects.filter(id=int(topic)).first()
        print (" ******* Found topic id *****")
        print (topic)

    if (reading_content):
    	subtopic1 = reading_content.subtopic1
    	topic = reading_content.reading_topic

    elif (subtopic1):
    	topic = subtopic1.topic




    a = (sheet.cell(row=row, column=6).value)
    if (a):
        a = getString(a)

    explanation = (sheet.cell(row=row, column=7).value)
    if (explanation):
        explanation = getString(explanation)




    t1 = (sheet.cell(row=row, column=13).value)
    if (t1):
        t1 = getString(t1)


    t2 = (sheet.cell(row=row, column=14).value)
    if (t2):
        t2 = getString(t2)

    t3 = (sheet.cell(row=row, column=15).value)
    if (t3):
        t3 = getString(t3)

        # t3= t3.encode('utf8')
    # t4 = (sheet.cell(row=row, column=13).value)
    # if (t4):
    #     t4 = getString(t4)

    q.choice_a = c1
    q.choice_b = c2
    q.choice_c = c3
    q.choice_d = c4


    q.mcq_answer = a
    q.explanation_text = explanation

    q.reading_content = reading_content
    q.subtopic1 = subtopic1
    q.reading_topic = topic

    q.tag1 = t1
    q.tag2 = t2
    q.tag3 = t3
    # q.tag5 = tag
    # q.update_date()
    if (not q.uploader):
        q.uploader = request.user

    # q.save()

    # print ('***********.....mcq has been uploaded and saved returning\n')
    # # print (q)
    # print ()

    # print ("******************** one mcq returning *************")
    # print (q)

    return q
























@transaction.atomic
def save_mcq(request, mcq_list, question_set):


	# return 1

	# print ("\n\n ******** mcq added to question set ********")

	cur_mcq = ""
	line = 10


	try:
		for mcq in mcq_list:
			cur_mcq = mcq
			mcq.save()
			mcq.update_date()
			line += 1

		# 	question_set.mcq_question.add(mcq)

		# question_set.save()
	except IntegrityError as e:
		messages.set_level(request, messages.ERROR)
		storage = messages.get_messages(request)
		# print (storage)
		messages.error(request, "There Is An Error At Line No: %s" % str(line))
		messages.error(request, "The Error MCQ Is %s " % cur_mcq)
		messages.error(request, e.message)
		# return e.message
		return 0



	return (line - 10)

		# render_to_response("template.html", {"message": e.message})






def create_quick_question_from_excel(request, sheet, row, reading_content):
    if ((sheet.cell(row=row, column=1).value) is None):
        # print ("in if part: " + str(sheet.cell(row=row, column=1).value))
        print ("************* NONE returning **********")
        return None
    # print ('***********.....mcq title is about to be been created \n')
    title = (sheet.cell(row=row, column=1).value)
    if (title):
        title = getString(title)
        title = title.decode('UTF8')
        q = Quick_Question(quick_question_text =title)
        q.content = reading_content

    answer = (sheet.cell(row=row, column=3).value)
    if (answer):
        answer = getString(answer)
        q.quick_question_answer = answer


    t1 = (sheet.cell(row=row, column=13).value)
    if (t1):
        t1 = getString(t1)


    t2 = (sheet.cell(row=row, column=14).value)
    if (t2):
        t2 = getString(t2)

    t3 = (sheet.cell(row=row, column=15).value)
    if (t3):
        t3 = getString(t3)



    q.tag1 = t1
    q.tag2 = t2
    q.tag3 = t3
    # q.tag5 = tag
    # q.update_date()
    if (not q.uploader):
        q.uploader = request.user

    return q
























@transaction.atomic
def save_quick_question(request, mcq_list, question_set):


	# return 1

	# print ("\n\n ******** mcq added to question set ********")

	cur_mcq = ""
	line = 10


	try:
		for mcq in mcq_list:
			cur_mcq = mcq
			mcq.save()
			mcq.update_date()
			line += 1

		# 	question_set.mcq_question.add(mcq)

		# question_set.save()
	except IntegrityError as e:
		messages.set_level(request, messages.ERROR)
		storage = messages.get_messages(request)
		# print (storage)
		messages.error(request, "There Is An Error At Line No: %s" % str(line))
		messages.error(request, "The Error MCQ Is %s " % cur_mcq)
		messages.error(request, e.message)
		# return e.message
		return 0



	return (line - 10)

		# render_to_response("template.html", {"message": e.message})























def func():
	print ("Hello alamin")








